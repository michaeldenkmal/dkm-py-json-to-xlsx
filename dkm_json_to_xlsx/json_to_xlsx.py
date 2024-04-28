from typing import Dict

import dkm_lib_json.json_util
from dkm_lib_json import json_util
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def wrap_arr_json_in_root_key(json_in_fp:str, json_out_fp:str,root_key:str):
    json_str = json_util.load_json_from_file(json_in_fp)
    wrapped_json = f'{{ "{root_key}": ' + json_str + f'}}'
    json_util.save_json_to_file(json_out_fp, wrapped_json)


def write_header(ws:Worksheet,row_idx:int, row:Dict[str, any]):
    keys = row.keys()
    col = 1
    for key in keys:
        ws.cell(row_idx, col, key)
        col += 1


def write_row_date(ws:Worksheet, row_idx:int,row:Dict[str, any]):
    keys = row.keys()
    col = 1
    for key in keys:
        ws.cell(row_idx, col, f"{row[key]}")
        col += 1



#https://openpyxl.readthedocs.io/en/stable/
def convert_json_to_xlsx (json_file:str, xlsx_out_path:str, arr_prop_name:str):
    json_str = dkm_lib_json.json_util.load_json_from_file(json_file)
    data_dict = dkm_lib_json.json_util.replace_dkm_date_strs(dkm_lib_json.json_util.json_2_dict(json_str))
    arr_rows = data_dict[arr_prop_name]
    wb = Workbook()
    ws = wb.active
    row_idx =1
    for row in arr_rows:
        if row_idx==1:
            write_header(ws, 1, row)
            row_idx +=1
        write_row_date(ws,row_idx, row)
        row_idx += 1
    wb.save(xlsx_out_path)




